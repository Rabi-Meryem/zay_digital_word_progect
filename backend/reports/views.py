from django.shortcuts import render

# Create your views here.
import csv
import io

from django.core.files.base import ContentFile
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from users.permissions import IsAdminOrSupervisor
from .models import Report
from .serializers import ReportSerializer
from .services import get_report_data, parse_date_param


class GenerateReportView(APIView):
    """
    GET /api/reports/generate/?format=pdf|excel|csv
        &date_from=2026-07-01&date_to=2026-07-14
        &agent_id=3&priority=HIGH&status=RESOLVED

    PDF et Excel sont persistés dans la table `reports` (fichier + métadonnées).
    CSV est un export ponctuel non persisté (ReportType ne supporte pas CSV).
    """
    permission_classes = [IsAuthenticated, IsAdminOrSupervisor]

    def get(self, request):
        print("========== GenerateReportView exécutée ==========")
        fmt = request.GET.get('exportFormat', 'pdf').lower()

        today = timezone.now().date()
        period_start = parse_date_param(request.GET.get('date_from'), today.replace(day=1))
        period_end = parse_date_param(request.GET.get('date_to'), today)

        data = get_report_data(
            date_from=period_start,
            date_to=period_end,
            agent_id=request.GET.get('agent_id'),
            priority=request.GET.get('priority'),
            ticket_status=request.GET.get('status'),
        )

        if fmt == 'excel':
            return self._excel_response(request, data, period_start, period_end)
        if fmt == 'csv':
            return self._csv_response(data)
        return self._pdf_response(request, data, period_start, period_end)

    # ── PDF ──────────────────────────────────────────────────────────────
    def _build_pdf_bytes(self, data):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Rapport de performance — ZAY Digital World", styles['Title']))
        elements.append(Paragraph(
            f"Généré le {data['generated_at'].strftime('%d/%m/%Y %H:%M')}", styles['Normal']
        ))
        elements.append(Spacer(1, 16))

        elements.append(Paragraph("Synthèse des KPIs", styles['Heading2']))
        k = data['kpis']
        kpi_table = Table([
            ['Total tickets', 'Ouverts', 'En cours', 'Résolus', 'Conformité SLA', 'Satisfaction'],
            [k['total'], k['open'], k['in_progress'], k['resolved'], f"{k['sla_compliance']}%", f"{k['satisfaction']}/5"],
        ])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 16))

        elements.append(Paragraph("Performance par agent", styles['Heading2']))
        agent_data = [['Agent', 'Traités', 'Conformité SLA', 'Satisfaction']]
        for a in data['agents']:
            agent_data.append([a['name'], a['handled'], f"{a['sla_compliance']}%", f"{a['satisfaction']}/5"])
        agent_table = Table(agent_data)
        agent_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D6A9F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(agent_table)
        elements.append(Spacer(1, 16))

        elements.append(Paragraph("Tickets escaladés", styles['Heading2']))
        esc_data = [['Ticket', 'Titre', 'Type', 'Date', 'Résolue']]
        for e in data['escalations']:
            esc_data.append([
                e['ticket_number'], e['title'][:40], e['type'],
                e['date'].strftime('%d/%m/%Y'), 'Oui' if e['resolved'] else 'Non',
            ])
        esc_table = Table(esc_data)
        esc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8A020')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(esc_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    def _pdf_response(self, request, data, period_start, period_end):
        pdf_bytes = self._build_pdf_bytes(data)

        report = Report.objects.create(
            generated_by=request.user,
            report_type=Report.ReportType.PDF,
            period_start=period_start,
            period_end=period_end,
        )
        report.file.save(f"rapport_{report.id}.pdf", ContentFile(pdf_bytes), save=True)

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_{report.id}.pdf"'
        return response

    # ── Excel ────────────────────────────────────────────────────────────
    def _build_excel_bytes(self, data):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        ws1 = wb.active
        ws1.title = 'KPIs'
        k = data['kpis']
        ws1.append(['Total', 'Ouverts', 'En cours', 'Résolus', 'Conformité SLA (%)', 'Satisfaction (/5)'])
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws1.append([k['total'], k['open'], k['in_progress'], k['resolved'], k['sla_compliance'], k['satisfaction']])

        ws2 = wb.create_sheet('Performance agents')
        ws2.append(['Agent', 'Traités', 'Conformité SLA (%)', 'Satisfaction (/5)'])
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
        for a in data['agents']:
            ws2.append([a['name'], a['handled'], a['sla_compliance'], a['satisfaction']])

        ws3 = wb.create_sheet('Escalades')
        ws3.append(['Ticket', 'Titre', 'Type', 'Date', 'Résolue'])
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
        for e in data['escalations']:
            ws3.append([
                e['ticket_number'], e['title'], e['type'],
                e['date'].strftime('%d/%m/%Y %H:%M'), 'Oui' if e['resolved'] else 'Non',
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _excel_response(self, request, data, period_start, period_end):
        excel_bytes = self._build_excel_bytes(data)

        report = Report.objects.create(
            generated_by=request.user,
            report_type=Report.ReportType.EXCEL,
            period_start=period_start,
            period_end=period_end,
        )
        report.file.save(f"rapport_{report.id}.xlsx", ContentFile(excel_bytes), save=True)

        response = HttpResponse(
            excel_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="rapport_{report.id}.xlsx"'
        return response

    # ── CSV (non persisté — ReportType ne le supporte pas) ────────────────
    def _csv_response(self, data):
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        writer.writerow(['--- KPIs ---'])
        writer.writerow(['Total', 'Ouverts', 'En cours', 'Résolus', 'Conformité SLA (%)', 'Satisfaction (/5)'])
        k = data['kpis']
        writer.writerow([k['total'], k['open'], k['in_progress'], k['resolved'], k['sla_compliance'], k['satisfaction']])
        writer.writerow([])

        writer.writerow(['--- Performance agents ---'])
        writer.writerow(['Agent', 'Traités', 'Conformité SLA (%)', 'Satisfaction (/5)'])
        for a in data['agents']:
            writer.writerow([a['name'], a['handled'], a['sla_compliance'], a['satisfaction']])
        writer.writerow([])

        writer.writerow(['--- Escalades ---'])
        writer.writerow(['Ticket', 'Titre', 'Type', 'Date', 'Résolue'])
        for e in data['escalations']:
            writer.writerow([
                e['ticket_number'], e['title'], e['type'],
                e['date'].strftime('%d/%m/%Y %H:%M'), 'Oui' if e['resolved'] else 'Non',
            ])

        response = HttpResponse(buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="rapport_zay.csv"'
        return response


class ReportListView(APIView):
    """GET /api/reports/ — historique des rapports PDF/Excel déjà générés."""
    permission_classes = [IsAuthenticated, IsAdminOrSupervisor]

    def get(self, request):
        reports = Report.objects.select_related('generated_by').order_by('-generated_at')[:50]
        return Response(ReportSerializer(reports, many=True, context={'request': request}).data)